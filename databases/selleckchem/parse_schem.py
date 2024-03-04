#!/usr/bin/env python3

# Interesting cases:
# - Multiple xlsx files, lots of duplicates across them
# - Some columns named differently across sheets
# - "Name" field sometimes contains one or more synonyms in ()
# - Also has a separate synonyms field that sometimes overlaps
# - Some names have () in them as part of the molecule name
# - Some columns are not present in all files (no synonyms in some)
# - Some extra rows/columns are formatted and so present in the data as None
# - Express-Pick Library (by Pfizer) has no name, not useful
# - Synonyms can be separated by any of ,;| or ', '.
#    in the case of comma without a space, the parse is ambiguous.
# - In a few spots they include the smiles code as a synonym

from __future__ import print_function
from atomicwrites import atomic_write
from openpyxl import Workbook, load_workbook

import logging
logger = logging.getLogger(__name__)

def sheet_values(sheet):
    out = []
    for row in sheet.rows:
        row_values = [cell.value for cell in row]
        out.append(row_values)
    logger.info("Loaded %d rows", len(out))
    return out


def index_of(where, what):
    if isinstance(what, str):
        return index_of(where, [what])

    where = [x.lower() for x in where if x]
    for inst in what:
        try:
            return where.index(inst.lower())
        except ValueError:
            pass
    raise ValueError('None of %s in list' % what)

import re
def parse_name(name):
    # Can match either "Chem Name"  or "Chem Name (Synonym)"
    # Also handles weird stuff like (-)-Epigenl... or blah(blah)blah...
    # returns name, syn
    m = re.match(r'^(.*?)(\([^(]*\))?$', name)
    assert m != None, "Unexpected name " + name
    g = m.groups()
    return g[0].strip(), g[1].strip(' ()') if g[1] else None


def split_synonyms(x):
    # Synonyms are split with a variety of different formats.
    # Prefer newlines, then semi-colons, then commas.
    # Commas are the worst because some names actually have commas.
    x = x.strip()
    if '\n' in x:
        sep_char = '\n'
    elif ';' in x:
        sep_char = ';'
    elif '|' in x:
        sep_char = '|'
    elif ', ' in x:
        sep_char = ', '
    else:
        # Our option of last resort, split everything on commas.
        sep_char = ','
    syns = [v.strip() for v in x.split(sep_char) if v.strip()]
    if len(syns) == 1 or sep_char != ',':
        return syns
    # If the seperator was a comma, it's an ambiguous parse, let's add
    # some length heuristics.
    return [syn for syn in syns if len(syn) >= 3]

def parse(input_file, seen):
    logger.info("Loading %s" % input_file)
    # read-only mode is wildly faster than the default.
    wb = load_workbook(input_file, read_only=True)
    logger.info("Workbook loaded, parsing")
    sheets = wb.sheetnames
    # Each one has a different name, but it's always the second sheet.
    info_sheet = wb[sheets[1]]
    data = sheet_values(info_sheet)
    try:
        id_col = index_of(data[0], ['cat', 'Catalog Number'])
        name_col = index_of(data[0], ['name', 'Product Name'])
    except ValueError:
        print("Skipping %s, no identifying information" % input_file)
        return []



    def synonyms(syn_col, row):
        syns = set()
        name, syn_from_name = parse_name(row[name_col])
        if syn_from_name:
            syns.update(split_synonyms(syn_from_name))

        if syn_col and syn_col != 'N/A':
            syns.update(split_synonyms(syn_col))

        return [('synonym', x.strip()) for x in sorted(syns) if x.strip()]

    def urls(url_col, row):
        if not url_col or url_col.endswith('/.html'):
            return []
        else:
            return [('url', url_col)]

    def smiles_code(smiles_col, row):
        if not smiles_col or not smiles_col.strip():
            return []
        return [('smiles_code', smiles_col.split(' ')[0])]

    col_map = [
            (['name', 'Product Name'], lambda x, row: [('canonical', parse_name(x)[0])]),
            (['CAS Number', 'CAS'], 'cas'),
            (['Synonyms'], synonyms),
            (['SMILES'], smiles_code),
            # Not bothering to extract for now, we can still construct a
            # search URL via ID.
            # (['URL'], urls),
            ]

    all_known_missing = {
            'Fragment-Library': ['Synonyms', 'URL'],
            }
    known_missing = set()
    for fileid, missings in list(all_known_missing.items()):
        if fileid in input_file:
            known_missing = set(missings)

    out = []
    for row in data[1:]:
        if len(row) <= id_col or not row[id_col]:
            continue
        id = 'SLK' + row[id_col]
        if id in seen:
            continue
        seen.add(id)
        for col_name, out_type in col_map:
            try:
                idx = index_of(data[0], col_name)
            except ValueError:
                if col_name[0] in known_missing:
                    continue
                else:
                    raise
            if isinstance(out_type, str):
                out_fn = lambda x, row: [(out_type, x)]
            else:
                out_fn = out_type

            val = row[idx]
            for attr, attrval in out_fn(val, row):
                attrval = str(attrval)
                if attrval.strip() != '':
                    out.append((id, attr, attrval.strip()))

    print("Adding %d lines to file, up to %d total IDs" % (len(out), len(seen)))
    return out

def parse_all(input_files):
    seen = set()
    out = [['selleckchem_id', 'attribute', 'value']]
    for input_file in input_files:
        out.extend(parse(input_file, seen))
    return out


def run(input_files, output_file):
    data = parse_all(input_files)
    with atomic_write(output_file, overwrite=True) as f:
        for row in data:
            f.write('\t'.join(row) + "\n")

if __name__ == "__main__":
    import argparse
    arguments = argparse.ArgumentParser(description="Parses out a selleckchem xlsx file.")

    arguments.add_argument('-o', '--output', help="Where to write the output")
    arguments.add_argument('inputs', nargs='+', help="Input files")

    args = arguments.parse_args()
    from dtk.log_setup import setupLogging
    setupLogging()

    run(args.inputs, args.output)

